#!/usr/bin/python3

import time
import json
import subprocess
import argparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def list_waf_policy_settings(subscription, resource_group, policy_name, timestamp):
    """列出WAF设置并保存为JSON文件"""
    ret_name = "settings_{}_{}.json".format(policy_name, timestamp)
    cmd = "az network application-gateway waf-policy policy-setting list --subscription {} --resource-group {} --policy-name {} > {}".format(subscription, resource_group, policy_name, ret_name)
    try:
        subprocess.run(cmd, shell=True, check=True, capture_output=True, encoding="utf-8")
    except Exception as e:
        print(e)
    return ret_name

def list_waf_policy_rules(subscription, resource_group, policy_name, timestamp):
    """列出WAF策略规则并保存为JSON文件"""
    ret_name = "rules_{}_{}.json".format(policy_name, timestamp)
    cmd = "az network application-gateway waf-policy custom-rule list --subscription {} --resource-group {} --policy-name {} > {}".format(subscription, resource_group, policy_name, ret_name)
    try:
        subprocess.run(cmd, shell=True, check=True, capture_output=True, encoding="utf-8")
    except Exception as e:
        print(e)
    return ret_name

def parse_waf_policy_settings(policy_settings_path):
    """解析WAF策略设置JSON字符串"""
    policy_settings_config = {}
    try:
        with open(policy_settings_path, 'r') as f:
            # 解析JSON文件
            data = json.loads(f.read())
        # 检查输入结构 - 应该是字典结构
        if isinstance(data, dict):
            policy_settings_config = data
    except json.JSONDecodeError as e:
        print(e)
        
    return policy_settings_config

def parse_waf_policy_rules(policy_rules_path):
    """解析WAF策略规则JSON字符串"""
    try:
        policy_rules_config = []
        with open(policy_rules_path, 'r') as f:
            # 解析JSON文件
            data = json.loads(f.read())
        # 检查输入结构 - 应该是列表
        if isinstance(data, list):
            policy_rules_config = data
    except json.JSONDecodeError as e:
        print(e)
    
    return policy_rules_config

def create_waf_excel(settings_config, rules_config, output_file='waf_policy.xlsx'):
    """创建WAF策略Excel文件"""
    
    # 创建Excel写入器
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # ====================
        # 1. 创建策略配置工作表
        # ====================
        if settings_config:
            # 将策略配置转换为DataFrame
            policy_df = pd.DataFrame(
                [(key, str(value)) for key, value in settings_config.items()],
                columns=['配置项', '值']
            )
            
            # 写入Excel
            policy_df.to_excel(writer, sheet_name='策略配置', index=False)
            
            # 获取工作表进行格式设置
            workbook = writer.book
            ws_policy = writer.sheets['策略配置']
            
            # 设置列宽
            ws_policy.column_dimensions['A'].width = 30
            ws_policy.column_dimensions['B'].width = 40
            
            # 设置标题样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws_policy[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        # ====================
        # 2. 创建规则列表工作表
        # ====================
        if rules_config:
            # 处理规则数据
            rules_data = []
            for rule in rules_config:
                # 提取基本信息
                rule_info = {
                    '名称': rule.get('name', ''),
                    '优先级': rule.get('priority', ''),
                    '动作': rule.get('action', ''),
                    '规则类型': rule.get('ruleType', ''),
                    '状态': rule.get('state', '')
                }
                
                # 处理匹配条件
                match_conditions = rule.get('matchConditions', [])
                if match_conditions:
                    for i, condition in enumerate(match_conditions):
                        prefix = f'条件{i+1}_' if len(match_conditions) > 1 else ''
                        
                        # 匹配变量
                        match_vars = condition.get('matchVariables', [])
                        if match_vars:
                            var_names = ', '.join([var.get('variableName', '') for var in match_vars])
                            rule_info[f'{prefix}匹配变量'] = var_names
                        
                        # 操作符
                        rule_info[f'{prefix}操作符'] = condition.get('operator', '')
                        
                        # 匹配值
                        match_values = condition.get('matchValues', [])
                        if match_values:
                            values_str = ', '.join(match_values)
                            rule_info[f'{prefix}匹配值'] = values_str
                        
                        # 否定条件
                        rule_info[f'{prefix}否定条件'] = '是' if condition.get('negationConditon', False) else '否'
                        
                        # 转换
                        transforms = condition.get('transforms', [])
                        if transforms:
                            rule_info[f'{prefix}转换'] = ', '.join(transforms)
                else:
                    rule_info['匹配条件'] = '无'
                
                rules_data.append(rule_info)
            
            # 创建DataFrame
            rules_df = pd.DataFrame(rules_data)
            
            # 写入Excel
            rules_df.to_excel(writer, sheet_name='规则列表', index=False)
            
            # 获取工作表进行格式设置
            ws_rules = writer.sheets['规则列表']
            
            # 设置列宽
            for col in ws_rules.columns:
                max_length = 0
                column = col[0].column_letter  # 获取列字母
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_rules.column_dimensions[column].width = adjusted_width
            
            # 设置标题样式
            for cell in ws_rules[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # 为规则行添加交替颜色
            light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for row in range(2, len(rules_df) + 2):
                if row % 2 == 0:
                    for cell in ws_rules[row]:
                        cell.fill = light_fill
        
        # ====================
        # 3. 创建摘要工作表
        # ====================
        summary_data = [
            ['WAF策略摘要', ''],
            ['生成时间', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['配置项统计', ''],
            ['策略配置数量', len(settings_config) if settings_config else 0],
            ['规则数量', len(rules_config) if rules_config else 0],
            ['', ''],
            ['策略模式', settings_config.get('mode', '未配置') if settings_config else '未配置'],
            ['WAF状态', settings_config.get('state', '未配置') if settings_config else '未配置'],
            ['', ''],
            ['规则动作统计', '']
        ]
        
        # 统计规则动作
        if rules_config:
            action_counts = {}
            for rule in rules_config:
                action = rule.get('action', '未指定')
                action_counts[action] = action_counts.get(action, 0) + 1
            
            for action, count in action_counts.items():
                summary_data.append([f'{action}规则', count])
        
        summary_df = pd.DataFrame(summary_data, columns=['项目', '值'])
        
        # 写入Excel
        summary_df.to_excel(writer, sheet_name='摘要', index=False)
        
        # 获取工作表进行格式设置
        ws_summary = writer.sheets['摘要']
        
        # 设置列宽
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # 设置标题样式
        title_font = Font(size=14, bold=True, color="366092")
        ws_summary['A1'].font = title_font
        
        # 设置其他标题
        for row in [1, 4, 8, 11]:
            ws_summary[f'A{row}'].font = Font(bold=True)
        
        print(f"Excel文件已生成: {output_file}")

def main():
    # GLOBAL VAR
    subscription = ""
    resource_group = ""
    policy_name = ""

    parser = argparse.ArgumentParser(
        description=
        """
        Cheatsheet:
        python az-list-waf-policy.py -s <subscription> -r <resource-group> -p <waf-policy-name>
        """
    )
    parser.add_argument("-s", "--subscription", type=str, help="waf policy subscription")
    parser.add_argument("-r", "--resource_group", type=str, help="waf policy resource group")
    parser.add_argument("-p", "--policy_name", type=str, help="waf policy name")
    arg = parser.parse_args()
    
    # 设置参数
    subscription = arg.subscription if arg.subscription else subscription
    resource_group = arg.resource_group if arg.resource_group else resource_group
    policy_name = arg.policy_name if arg.policy_name else policy_name

    # 获取时间戳
    timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())

    # 解析WAF策略
    settings_config = parse_waf_policy_settings(
        list_waf_policy_settings(subscription, resource_group, policy_name, timestamp)
    )
    rules_config = parse_waf_policy_rules(
        list_waf_policy_rules(subscription, resource_group, policy_name, timestamp)
    )

    # print(settings_config)
    # print(rules_config)

    # 创建Excel文件
    create_waf_excel(settings_config, rules_config, 'waf_policy_{}_{}.xlsx'.format(policy_name, timestamp))

if __name__ == "__main__":
    main()
