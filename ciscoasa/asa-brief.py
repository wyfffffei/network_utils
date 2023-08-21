#/usr/local/bin/python3
import re
import time
import ipaddress

from ciscoconfparse import CiscoConfParse

# pip3 install ciscoconfparse
# https://github.com/mpenning/ciscoconfparse
# http://pennington.net/tutorial/ciscoconfparse/ccp_tutorial.html

# GLOBAL VAR
_acl_column = ["源端口", "ACL类型", "动作", "协议", "源地址", "目标地址", "端口"]
_obj_column = ["对象(组)名", "映射IP/段"]
_intf_column = ["interface name", "nameif", "security-level", "ip address", "net mask", "vlan", "shutdown", "with group", "description"]
_route_column = ["IP", "Netmask", "Next hop"]


class ASA_Obj:
    def __init__(self, name, type, group=0):
        self.name = name
        self.type = type
        self.group = group
        self.children = []

    @staticmethod
    def search_objects(config) -> dict:
        if not isinstance(config, CiscoConfParse):
            # print("CiscoConfParse Object is required.")
            return None

        obj_dict = dict()
        for obj in config.find_objects('^object'):
            obj_line = obj.text.split(" ")
            # object or object-group xxx
            isgroup = 0 if obj_line[0] == "object" else 1
            # 创建 ASA 自定义对象
            asa_obj = ASA_Obj(' '.join(obj_line[2:]), obj_line[1], isgroup)
            if not obj.children:
                continue
            for item in obj.children:
                asa_obj.add_child(item.text.strip())
            if not obj_dict.get(str(asa_obj)):
                obj_dict[str(asa_obj)] = []
            obj_dict[str(asa_obj)].append(asa_obj)
        return obj_dict

    def add_child(self, obj):
        self.children.append(obj)

    def rm_child(self, obj):
        if obj in self.children:
            self.children.pop(obj)

    def __str__(self) -> str:
        return self.name


def search_hostname(config):
    hostname = config.find_objects("^hostname")
    if not hostname:
        return None
    return hostname[0].text[9:]


def search_access_rule(config):
    # global _acl_column

    acl_rules = config.find_objects("^access-list")
    obj_all = ASA_Obj.search_objects(config)
    for acl in acl_rules:
        acl = acl.text.split(" ")
        if acl[2] == "remark":
            yield (acl[:3] + [' '.join(acl[3:])])

        prefix = acl[1:5]           # 接口、策略类型、行为、协议
        acl = acl[5:]               # 访问源、访问目的、访问端口

        resv = []
        skip = False
        addr_block = ""
        # 解析 ->（访问源、访问目的、访问端口）
        for eid, entry in enumerate(acl):
            if skip:
                skip = False
                continue
            if entry == "any":
                resv.append(entry)
            elif entry == "host" or entry == "eq":
                resv.append(acl[eid + 1])
                skip = True
            elif entry == "object" or entry == "object-group":
                obj = acl[eid + 1]
                if obj_all.get(obj):
                    # 添加对象超链接
                    obj = "=HYPERLINK(\"#对象表!a{}\",\"{}\")".format(list(obj_all.keys()).index(obj) + 2, obj)
                resv.append(obj)
                skip = True
            elif entry == "log":
                continue
            else:
                # addr block
                if not addr_block:
                    addr_block = entry
                    continue
                try:
                    addr_block += ('/' + str(ipaddress.ip_network(addr_block + "/" + entry).prefixlen))
                except Exception as e:
                    addr_block += (" " + entry)
                resv.append(addr_block)
                addr_block = ""

        yield (prefix + resv)


def search_interface(config):
    # global _intf_column

    acl_gs = config.find_objects("^access-group\s")
    AG = dict()
    for acl_g in acl_gs:
        acl_g = acl_g.text.split(" ")
        AG[acl_g[-1]] = acl_g[1]

    intf_obj = config.find_objects('^interface\s+')
    for obj in intf_obj:
        intf_name = ' '.join(obj.text.split(" ")[1:])                       # 接口名

        intf_nameif = obj.re_match_iter_typed(                              # nameif
            r'^\snameif\s(\S+)', result_type=str, group=1, default='')
        
        intf_sec = obj.re_match_iter_typed(                                 # 优先级
            r'^\ssecurity-level\s(\d+)', result_type=str, group=1, default='')
        
        ip_addr = obj.re_match_iter_typed(                                  # IP地址
            r'^\sip\saddress\s(\d+\.\d+\.\d+\.\d+)\s(\d+\.\d+\.\d+\.\d+)', result_type=str, group=1, default='')
        
        net_mask = obj.re_match_iter_typed(                                 # 子网掩码
            r'^\sip\saddress\s(\d+\.\d+\.\d+\.\d+)\s(\d+\.\d+\.\d+\.\d+)', result_type=str, group=2, default='')
        
        vlan = obj.re_match_iter_typed(                                     # vlan号
            r'^\svlan\s(\d+)', result_type=str, group=1, default='')
        
        shutdown = obj.re_match_iter_typed(                                 # 开关状态
            r'^\s(shutdown)', result_type=str, group=1, default='')
        
        acl_group = AG.get(intf_nameif) if AG.get(intf_nameif) else ""      # 关联 ACL-Group

        description = obj.re_match_iter_typed(                              # 描述
            r'^\sdescription\s(.*)', result_type=str, group=1, default='')
        
        yield [intf_name, intf_nameif, intf_sec, ip_addr, net_mask, vlan, shutdown, acl_group, description]


def search_route_table(config):
    # global _route_column

    route_obj = config.find_objects('^route')
    for obj in route_obj:
        route_table = re.findall(r'\d+.\d+.\d+.\d+', obj.text)
        if not isinstance(route_table, list):
            continue
        yield route_table


def output_2_excel(config):
    global _acl_column, _obj_column, _intf_column, _route_column

    from openpyxl import Workbook
    wb = Workbook()

    ws = wb.active
    ws.title = "Cisco ASA 防火墙 ACL 访问策略"
    ws.append(_acl_column)
    for line in search_access_rule(config):
        ws.append(line)

    ws2 = wb.create_sheet(title="对象表")
    ws2.append(_obj_column)
    for o_name, obj in ASA_Obj.search_objects(config).items():
        obj_line = ""
        for item in obj:
            obj_line += '\n'.join(item.children) + '\n'
        ws2.append([o_name, obj_line.strip()])

    ws3 = wb.create_sheet(title="接口信息")
    ws3.append(_intf_column)
    for interface in search_interface(config):
        ws3.append(interface)

    ws4 = wb.create_sheet(title="路由表")
    ws4.append(_route_column)
    for route in search_route_table(config):
        ws4.append(route)

    hostname = search_hostname(config)
    wb.save(hostname + time.strftime("-%Y%m%d", time.localtime()) + ".xlsx")


def main():
    import sys
    if not sys.argv[1]:
        print("请添加参数: asa配置文件路径")
        sys.exit(-1)
    path = sys.argv[1]

    # GLOBAL: 配置源文件路径
    # path = "conf/dmz-asa.txt"

    config = CiscoConfParse(path, syntax="asa")
    output_2_excel(config)


if __name__ == "__main__":
    main()

