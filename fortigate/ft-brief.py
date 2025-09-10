#!/usr/local/bin/python3
import os
import re
import json
import time
import ipaddress

# 列出 飞塔 防火墙的相关配置
# python3 ft-brief.py <ft_config_path>

def json2dict(path):
    if not os.path.exists(path) or path[-5:] != ".json":
        print("File path is invalid.")
        return None

    with open(path, 'r') as f:
        return json.loads(f.read())


def subnet_caculator(ip):
    # 计算 IP 地址段包含范围
    if " " in ip:
        ip = re.sub(" ", "/", ip)
    try:
        net = ipaddress.ip_interface(ip).network
        if str(net.netmask) == "255.255.255.255":
            return str(net)[:-3]
        if str(net.netmask) == "255.255.255.254":
            return str(net[0]) + '-' + str(net[-1])
        return str(net[1]) + '-' + str(net[-2])
    except Exception as e:
        return ""


def search_hostname(config):
    # 查找 hostname
    _config_global = config["CONFIG"]["config system global"]

    host_pat = r"hostname\s\"\S+\""
    hostname = re.search(host_pat, ''.join(_config_global))
    return hostname.group().split("\"")[-2] if hostname else None


def search_interface(config):
    # 查找 接口信息
    _interface_all = config["CONFIG"]["config system interface"]

    for intf in _interface_all:
        intf_name = ''.join(list(intf.keys())[0]).split(" ")[-1].strip("\"")        # 接口名称
        intf = ' '.join(list(intf.values())[0])                                     # 接口配置原始信息
        
        ip_pat = r"\d+\.\d+\.\d+\.\d+\s\d+\.\d+\.\d+\.\d+"                          # 正则匹配IP地址段
        root_int_pat = r"set\sinterface\s\"\S+\""                                   # 正则匹配父接口
        alias_pat = r"set\salias\s\"\S+\""                                          # 正则匹配接口别名
        vlanid_pat = r"set\svlanid\s\d+"                                            # 正则匹配vlanid

        if re.search(ip_pat, intf):
            if re.search(root_int_pat, intf):
                root_int = re.search(root_int_pat, intf).group().split(" ")[-1].strip("\"")
            else:
                root_int = intf_name
                intf_name = ""
            intf_alias = re.search(alias_pat, intf).group().split(" ")[-1].strip("\"") if re.search(alias_pat, intf) else ""
            ip_net = re.search(ip_pat, intf).group()
            vlanid = re.search(vlanid_pat, intf).group().split(" ")[-1] if re.search(vlanid_pat, intf) else ""
            # return ->（接口、子接口、接口别名、vlanid、地址段、地址段范围）
            yield (root_int, intf_name, intf_alias, vlanid, ip_net, subnet_caculator(ip_net))


def search_ft_object(config):
    # 查找 配置对象
    try:
        _addr = config["CONFIG"]["config firewall address"]
    except Exception as e:
        print("No Address found.")
        return []

    for addr in _addr:
        addr_name = list(addr.keys())[0].split("\"")[-2]
        addr_value = ' '.join(list(addr.values())[0])

        subnet_pat = r"set\ssubnet\s\d+\.\d+\.\d+\.\d+\s\d+\.\d+\.\d+\.\d+"
        type_pat = r"set\stype\s\S+"
        fqdn_pat = r"set\sfqdn\s\"\S+\""
        intf_pat = r"set\sinterface\s\"\S+\""
        asso_intf_pat = r"set\sassociated-interface\s\"\S+\""
        start_ip_pat = r"set\sstart-ip\s\d+\.\d+\.\d+\.\d+"
        end_ip_pat = r"set\send-ip\s\d+\.\d+\.\d+\.\d+"

        subnet = re.search(subnet_pat, addr_value).group()[11:] if re.search(subnet_pat, addr_value) else "-"
        subnet_range = subnet_caculator(subnet) if subnet_caculator(subnet) else "-"
        type = re.search(type_pat, addr_value).group().split(" ")[-1] if re.search(type_pat, addr_value) else "-"
        fqdn = re.search(fqdn_pat, addr_value).group().split(" ")[-1][1:-1] if re.search(fqdn_pat, addr_value) else "-"
        intf = re.search(intf_pat, addr_value).group().split(" ")[-1][1:-1] if re.search(intf_pat, addr_value) else "-"
        asso_intf = re.search(asso_intf_pat, addr_value).group().split(" ")[-1][1:-1] if re.search(asso_intf_pat, addr_value) else "-"
        start_ip = re.search(start_ip_pat, addr_value).group().split(" ")[-1] if re.search(start_ip_pat, addr_value) else "-"
        end_ip = re.search(end_ip_pat, addr_value).group().split(" ")[-1] if re.search(end_ip_pat, addr_value) else "-"
        # 提取 comment
        comment = "-"
        for iid, item in enumerate(addr_value.split("\"")):
            if "comment " in item:
                comment = addr_value.split("\"")[iid + 1]
                break
        yield (addr_name, subnet, subnet_range, type, fqdn, intf, asso_intf, start_ip, end_ip, comment)


def search_ft_object_group(config):
    # 查找 配置对象组
    try:
        _addrgrp = config["CONFIG"]["config firewall addrgrp"]
    except Exception as e:
        print("No Address Group found.")
        return []

    for addrgrp in _addrgrp:
        addrgrp_name = list(addrgrp.keys())[0].split("\"")[-2]
        addrgrp_value = list(addrgrp.values())[0]

        for item in addrgrp_value:
            if "set member" in item:
                yield (addrgrp_name, '\n'.join([member.strip("\"") for member in item.split(" ")[2:]]))


def search_ipsecvpn(config):
    # 查找 ipsec vpn 兴趣流
    try:
        _phase1_config = config["CONFIG"]["config vpn ipsec phase1-interface"]
        _phase2_config = config["CONFIG"]["config vpn ipsec phase2-interface"]
    except Exception as e:
        print("No IPSec VPN found.")
        return ({}, "")

    phase1 = dict()
    for item in _phase1_config:
        ph1_name = list(item.keys())[0].split(" ")[-1].strip("\"")
        ph1_item = ' '.join(list(item.values())[0])
        ph1_int_pat = r"set\sinterface\s\"\S+\""                    # 匹配VPN接口
        ph1_remote_pat = r"set\sremote-gw\s\d+.\d+.\d+.\d+"         # 匹配对端地址

        ph1_int = re.search(ph1_int_pat, ph1_item).group()[14:].strip("\"") if re.search(ph1_int_pat, ph1_item) else ""
        ph1_remote = re.search(ph1_remote_pat, ph1_item).group()[14:] if re.search(ph1_remote_pat, ph1_item) else ""
        # return --> (VPN接口、对端IP)
        if not phase1.get(ph1_name):
            phase1[ph1_name] = (ph1_int, ph1_remote)

    phase2 = dict()
    phase2_pat = r"set\sphase1name\s\"\S+\""                                                    # 匹配感兴趣流名称
    src_name_pat = r"set\ssrc-name\s\"\S+\""                                                    # 匹配感兴趣流源地址
    dst_name_pat = r"set\sdst-name\s\"\S+\""                                                    # 匹配感兴趣流目标地址
    src_subnet_pat = r"set\ssrc-subnet\s\d+\.\d+\.\d+\.\d+\s\d+\.\d+\.\d+\.\d+"                 # 匹配感兴趣流源地址
    dst_subnet_pat = r"set\sdst-subnet\s\d+\.\d+\.\d+\.\d+\s\d+\.\d+\.\d+\.\d+"                 # 匹配感兴趣流目标地址

    for item in _phase2_config:
        ph2_item = list(item.keys())[0].split(" ")[-1].strip("\"")
        ph2_detail = ' '.join(list(item.values())[0])
        ph2_name = re.search(phase2_pat, ph2_detail)
        if not ph2_name:
            continue
        ph2_name = ph2_name.group().split(" ")[-1].strip("\"")

        # 从 src-name 和 src-subnet 中选出非空元素
        src = {re.search(src_name_pat, ph2_detail), re.search(src_subnet_pat, ph2_detail)} ^ {None,}
        src_interest = ' '.join(src.pop().group().split(" ")[2:]).strip("\"") if src else ""
        dst = {re.search(dst_name_pat, ph2_detail), re.search(dst_subnet_pat, ph2_detail)} ^ {None,}
        dst_interest = ' '.join(dst.pop().group().split(" ")[2:]).strip("\"") if dst else ""

        # return --> (兴趣流源地址、兴趣流源范围、兴趣流目标地址、兴趣流目标范围)
        if not phase2.get(ph2_name):
            phase2[ph2_name] = {ph2_item: (src_interest, subnet_caculator(src_interest), dst_interest, subnet_caculator(dst_interest))}
            continue
        phase2[ph2_name].update({ph2_item: (src_interest, subnet_caculator(src_interest), dst_interest, subnet_caculator(dst_interest))})
    return (phase1, phase2)


def search_static_route(config):
    # 查找 静态路由
    try:
        _static_route = config["CONFIG"]["config router static"]
    except Exception as e:
        print("No Static Route found.")
        return []

    dst_pat = r"set\sdst\s\d+\.\d+\.\d+\.\d+\s\d+\.\d+\.\d+\.\d+"   # 匹配目标段
    gateway_pat = r"set\sgateway\s\d+\.\d+\.\d+\.\d+"               # 匹配网关
    interface_pat = r"set\sdevice\s\"\S+\""                         # 匹配接口
    comment_pat = r"set\scomment\s\"\S+\""                          # 匹配备注
    priority_pat = r"set\spriority\s\d+"                            # 匹配优先级
    status_pat = r"set\sstatus\s\S+"                                # 匹配状态

    for route in _static_route:
        route_id = list(route.keys())[0][5:]                        # edit "<route_id>"
        route_item = ' '.join(list(route.values())[0])              # route source config

        dst = re.search(dst_pat, route_item).group()[8:] if re.search(dst_pat, route_item) else "0.0.0.0 0.0.0.0"
        gateway = re.search(gateway_pat, route_item).group()[12:] if re.search(gateway_pat, route_item) else ""
        interface = re.search(interface_pat, route_item).group()[11:].strip("\"") if re.search(interface_pat, route_item) else ""
        comment = re.search(comment_pat, route_item).group()[12:].strip("\"") if re.search(comment_pat, route_item) else ""
        priority = re.search(priority_pat, route_item).group()[13:] if re.search(priority_pat, route_item) else ""
        status = re.search(status_pat, route_item).group()[11:] if re.search(status_pat, route_item) else ""
        # return -> (路由ID、目标段、目标段范围、网关、接口、备注、优先级、状态)
        yield (route_id, dst, subnet_caculator(dst), gateway, interface, comment, priority, status)


def search_vip(config):
    try:
        _vip = config["CONFIG"]["config firewall vip"]
    except Exception as e:
        print("No Virtual IP Service found.")
        return []

    expip_pat = r"set\sextip\s\d+\.\d+\.\d+\.\d+"
    mappedip_pat = r"set\smappedip\s\"\d+\.\d+\.\d+\.\d+\""
    extport_pat = r"set\sextport\s\d+"
    mappedport_pat = r"set\smappedport\s\d+"

    for ext in _vip:
        service_name = list(ext.keys())[0].split(" ")[-1][1:-1]
        ext_config = ' '.join(list(ext.values())[0])

        expip = re.search(expip_pat, ext_config).group().split(" ")[-1] if re.search(expip_pat, ext_config) else ""
        mappedip = re.search(mappedip_pat, ext_config).group().split(" ")[-1][1:-1] if re.search(mappedip_pat, ext_config) else ""
        extport = re.search(extport_pat, ext_config).group().split(" ")[-1] if re.search(extport_pat, ext_config) else ""
        mappedport = re.search(mappedport_pat, ext_config).group().split(" ")[-1] if re.search(mappedport_pat, ext_config) else ""

        yield(service_name, expip, mappedip, extport, mappedport)


def output_2_excel(ft_policy):
    from openpyxl import Workbook
    wb = Workbook()

    # 创建表1（接口表）==> ["接口", "子接口", "接口别名", "vlanid", "地址段", "地址段范围"]
    ws_interface = wb.active
    ws_interface.title = "接口表"
    ws_interface.append(["接口", "子接口", "接口别名", "vlanid", "地址段", "地址段范围"])
    for intf in search_interface(ft_policy):
        ws_interface.append(intf)

    # 创建表2（IPSEC VPN表）==> ["IPSEC VPN名称", "VPN接口", "对端IP", "兴趣流名称", "兴趣流源地址", "兴趣流源范围", "兴趣流目标地址", "兴趣流目标范围"]
    ws_ipsec = wb.create_sheet("IPSEC VPN表")
    ws_ipsec.append(["IPSEC VPN名称", "VPN接口", "对端IP", "兴趣流名称", "兴趣流源地址", "兴趣流源范围", "兴趣流目标地址", "兴趣流目标范围"])
    ph1, ph2 = search_ipsecvpn(ft_policy)
    for ph_name, ph_item in ph1.items():
        for id_ph, (interest, stream_interest) in enumerate(ph2[ph_name].items()):
            if id_ph == 0:
                ws_ipsec.append((ph_name,) + ph_item + (interest,) + stream_interest)
                continue
            ws_ipsec.append(("", "", "", interest) + stream_interest)

    # 创建表3（路由表）==> ["路由ID", "目标段", "目标段范围", "网关", "接口", "备注", "优先级", "状态"]
    ws_route = wb.create_sheet("路由表")
    ws_route.append(["路由ID", "目标段", "目标段范围", "网关", "接口", "备注", "优先级", "状态"])
    for line in search_static_route(ft_policy):
        ws_route.append(line)

    # 创建表4（公网映射表）==> ["映射名称", "公网IP", "映射IP", "公网端口", "映射端口"]
    ws_vip = wb.create_sheet("公网映射表")
    ws_vip.append(["映射名称", "公网IP", "映射IP", "公网端口", "映射端口"])
    for line in search_vip(ft_policy):
        ws_vip.append(line)

    # 创建表5（对象表）==> ["对象名称", "子网", "类型", "fqdn", "接口", "关联接口", "起始IP", "结束IP"]
    ws_addr = wb.create_sheet("对象表")
    ws_addr.append(["对象名称", "子网", "子网范围", "类型", "fqdn", "接口", "关联接口", "起始IP", "结束IP", "备注"])
    for line in search_ft_object(ft_policy):
        ws_addr.append(line)

    # 创建表6（对象组表）==> ["对象组名称", "成员"]
    ws_addr = wb.create_sheet("对象组表")
    ws_addr.append(["对象组名称", "成员"])
    for line in search_ft_object_group(ft_policy):
        ws_addr.append(line)

    # 输出文件名格式：<hostname>-<timestamp>.xlsx
    wb.save(search_hostname(ft_policy) + "-brief" + time.strftime("-%Y%m%d", time.localtime()) + ".xlsx")


def main():
    import sys
    current_dir = os.path.dirname(os.path.abspath(__file__))
    sys.path.append(os.sep.join(current_dir.split(os.sep)[:-1]))

    from conf_parser import FortiGate

    # GLOBAL: 配置源文件路径
    path = "xxx/xxx.conf"

    if not sys.argv[1]:
        print("请添加参数: fortigate配置文件路径")
        sys.exit(-1)
    path = sys.argv[1]
    ft_policy = FortiGate(path).parse_policy()

    output_2_excel(ft_policy)


if __name__ == "__main__":
    main()
