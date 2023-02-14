import os
import json
import textwrap
from argparse import HelpFormatter


def json2dict(path):
    if not os.path.exists(path) or path[-5:] != ".json":
        print("File path is invalid.")
        return None

    with open(path, 'r') as f:
        return json.loads(f.read())


def dict2json(dict, output_path, encoding="utf-8"):
    with open(output_path, "w", encoding=encoding) as f:
        f.write(json.dumps(dict, indent=4, separators=(',', ':')))


def append_excel(path):
    # 追加创建 excel ，返回 sheet 对象
    try:
        from openpyxl import load_workbook, Workbook
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
        return wb

    except ImportError as e:
        print(e)
        return None
    
class RawFormatter(HelpFormatter):
    def _fill_text(self, text, width, indent):
        return "\n".join([textwrap.fill(line, width) for line in textwrap.indent(textwrap.dedent(text), indent).splitlines()])
